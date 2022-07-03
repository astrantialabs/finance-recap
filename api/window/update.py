import json
import os
import datetime
import timeit
import openpyxl
import gridfs

from pymongo import MongoClient
from openpyxl.styles import *
from win32com import client

class Excel():
    def __init__(self, file_path: str, active_sheet: int = 1):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)

        wb_sheet = self.workbook.sheetnames

        self.active_sheet = self.workbook[wb_sheet[active_sheet - 1]]


    def create_file(file_path: str):
        wb = openpyxl.Workbook()
        wb.save(file_path)


    def save(self):
        self.workbook.save(self.file_path)


    def create_sheet(self, new_sheet_name: str):
        self.workbook.create_sheet(new_sheet_name)


    def change_sheet(self, active_sheet: int or str):
        if(type(active_sheet) == int):
            self.active_sheet = self.workbook[self.workbook.sheetnames[active_sheet - 1]]

        elif(type(active_sheet) == str):
            self.active_sheet = self.workbook[active_sheet]


    def change_sheet_name(self, old_sheet_name, new_sheet_name):
        self.workbook[old_sheet_name].title = new_sheet_name
    

    def delete_sheet(self, deleted_sheet: int or str):
        if(type(deleted_sheet) == int):
            del self.workbook[self.workbook.sheetnames[deleted_sheet - 1]]

        elif(type(deleted_sheet) == str):
            del self.workbook[deleted_sheet]


    def adjust_width(self, start_range: any, end_range: any, extra_width: int = 0, width_limit: int = 0):
        start_column, start_row = Excel.convert_range(start_range)
        end_column, end_row = Excel.convert_range(end_range)

        width_dict = {}
        for column in range(start_column, end_column + 1):
            temp_width_array = []
            cell_column_letter =  self.active_sheet.cell(row = 1, column = column).column_letter

            for row in range(start_row, end_row + 1):
                temp_value = self.active_sheet.cell(row = row, column = column).value

                if(type(temp_value) in (int, str)):
                    temp_width_array.append(len(str(temp_value)))


            max_width_value = max(temp_width_array)

            if(width_limit > 0):
                if(max_width_value > width_limit):
                    max_width_value = width_limit

                    self.alignment_multiple([column, start_row], [column, end_row], wrap = True)


            width_dict[cell_column_letter] = max_width_value

        
        for column, width in width_dict.items():
            self.active_sheet.column_dimensions[column].width = (width + 1 + extra_width)


    def check_range(range: any):
        if (type(range) not in (str, list)):
            raise TypeError("Range must be a type of string or list")

        elif (type(range) == str):
            if (range.isalpha() or range.isnumeric()):
                raise TypeError("Range string must be a combination of character and number")

        elif (type(range) == list):
            if (len(range) == 2):
                for i in range:
                    if type(i) not in (str, int):
                        raise TypeError("Range list can only have a type of string and integer for its values")

            else:
                raise TypeError("Range list can only have 2 values")
                

    def convert_range(range: any):
        Excel.check_range(range)

        if (type(range) == str):
            column = Excel.check_and_convert_string_value(''.join(x for x in range if not x.isdigit()))
            row = int(''.join(x for x in range if x.isdigit()))

        elif (type(range) == list):
            if (type(range[0]) == str):
                column = Excel.check_and_convert_string_value(range[0])
            
            elif (type(range[0]) == int):
                column = range[0]

            if (type(range[1]) == str):
                row = Excel.check_and_convert_string_value(range[1])
            
            elif (type(range[1]) == int):
                row = range[1]

        return column, row


    def check_and_convert_string_value(value: any):
        if(type(value) == str):
            value = [ord(x) - 96 for x in value.lower()]

            new_value = 0
            for i in range(len(value)):
                new_value += value[i] * 26**(len(value) - (i + 1))

        return new_value

    
    def attributes_string(list_of_attributes: any):
        attributes_string = ""
        for i, attribute in enumerate(list_of_attributes):
            if(i == 0):
                attributes_string += attribute
            
            else:
                attributes_string += f", {attribute}"

        return attributes_string


    def change_cell_number_format_singular(self, range: any, number_format: str):
        column, row = Excel.convert_range(range)
        self.active_sheet.cell(row = row, column = column).number_format = number_format


    def change_cell_number_format_multiple(self, start_range: any, end_range: any, number_format: str):
        start_column, start_row = Excel.convert_range(start_range)
        end_column, end_row = Excel.convert_range(end_range)

        for row in range(start_row, end_row + 1):
            for column in range(start_column, end_column + 1):
                self.active_sheet.cell(row = row, column = column).number_format = number_format


    #region Get
    def get_value_singular(self, range: any):
        column, row = Excel.convert_range(range)
        value = self.active_sheet.cell(row = row, column = column).value

        return value

        
    def get_value_multiple(self, start_range: any, end_range: any):
        start_column, start_row = Excel.convert_range(start_range)
        end_column, end_row = Excel.convert_range(end_range)
        
        value = []
        for row in range(start_row, end_row + 1):
            for column in range(start_column, end_column + 1):
                temp_value = self.active_sheet.cell(row = row, column = column).value
                value.append(temp_value)

        return value


    def get_value_multiple_2d(self, start_range: any, end_range: any):
        start_column, start_row = Excel.convert_range(start_range)
        end_column, end_row = Excel.convert_range(end_range)

        value_array = []
        for row in range(start_row, end_row + 1):
            temp_value_array = []
            for column in range(start_column, end_column + 1):
                temp_value = self.active_sheet.cell(row = row, column = column).value
                temp_value_array.append(temp_value)
            
            value_array.append(temp_value_array)

        return value_array

    #endregion Get


    #region Write
    def write_value_singular(self, range: any, value: any):
        if(type(value) == list):
            raise TypeError("Use write_value_multiple function if the value type is a list")

        column, row = Excel.convert_range(range)
            
        self.active_sheet.cell(row = row, column = column, value = value)


    def write_value_multiple(self, start_range: any, end_range: any, value: any):
        start_column, start_row = Excel.convert_range(start_range)
        end_column, end_row = Excel.convert_range(end_range)
        
        if(type(value) == list):
            for check_value in value:
                if(type(check_value) == list):
                    raise TypeError("Use write_value_multiple_2d function if the value is a 2D list")


            value_counter = 0
            for row in range(start_row, end_row + 1):
                for column in range(start_column, end_column + 1):
                    self.active_sheet.cell(row = row, column = column, value = value[value_counter])
                    value_counter += 1


        elif(type(value) in (str, int, bool, float)):
            for row in range(start_row, end_row + 1):
                for column in range(start_column, end_column + 1):
                    self.active_sheet.cell(row = row, column = column, value = value)       


    def write_value_multiple_2d(self, start_range: any, value: any):
        if(type(value) == list):
            value_is_valid = True
            for check_value in value:
                if(type(check_value) != list):
                    value_is_valid = False


            if(value_is_valid):
                start_column, start_row = Excel.convert_range(start_range)
                end_column, end_row = start_column + len(value[0]), start_row + len(value)

                for x, row in enumerate(range(start_row, end_row)):
                    for y, column in enumerate(range(start_column, end_column)):
                        self.active_sheet.cell(row = row, column = column, value = value[x][y])


            elif(not value_is_valid):
                raise TypeError("Value must be a 2D list")
        
        else:
            raise TypeError("Value must be a 2D list")

    #endregion Write


    #region Merge & Unmerge
    def merge(self, start_range: any, end_range: any):
        start_column, start_row = Excel.convert_range(start_range)
        end_column, end_row = Excel.convert_range(end_range)

        self.active_sheet.merge_cells(start_row = start_row, start_column = start_column, end_row = end_row, end_column = end_column)

    
    def unmerge(self, start_range: any, end_range: any):
        start_column, start_row = Excel.convert_range(start_range)
        end_column, end_row = Excel.convert_range(end_range)

        self.active_sheet.unmerge_cells(start_row = start_row, start_column = start_column, end_row = end_row, end_column = end_column)

    #endregion Merge & Unmerge


    #region Font
    def font_attributes(**attributes: any):
        list_of_attributes = []
        if("font" in attributes):
            if(type(attributes.get("font")) == str):
                font_name = attributes.get("font")
                temp_attribute = f"name='{font_name}'"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Font data type needs to be a string")

        if("size" in attributes):
            if(type(attributes.get("size")) in (str, int)):
                font_size = int(attributes.get("size"))
                temp_attribute = f"size={font_size}"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Size data type needs to be a string or an integer")

        if("color" in attributes):
            if(type(attributes.get("color")) == str):
                color_name = attributes.get("color")
                temp_attribute = f"color='{color_name}'"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Color data type needs to be a string")

        if("underline" in attributes):
            if(type(attributes.get("underline")) == str):
                underline_name = attributes.get("underline")
                underline_name = (underline_name[0].lower() + underline_name[1:]).replace(" ", "")
                temp_attribute = f"underline='{underline_name}'"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Underline data type needs to be a string")

        if("bold" in attributes):
            if(type(attributes.get("bold")) == bool):
                is_bold = attributes.get("bold")
                temp_attribute = f"bold={is_bold}"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Bold data type needs to be a boolean")

        if("italic" in attributes):
            if(type(attributes.get("italic")) == bool):
                is_italic = attributes.get("italic")
                temp_attribute = f"italic={is_italic}"

                list_of_attributes.append(temp_attribute)
            
            else:
                raise TypeError("Italic data type needs to be a boolean")

        if("strike" in attributes):
            if(type(attributes.get("strike")) == bool):
                is_strike = attributes.get("strike")
                temp_attribute = f"strike={is_strike}"

                list_of_attributes.append(temp_attribute)
        
            else:
                raise TypeError("Strike data type needs to be a boolean")

        return Excel.attributes_string(list_of_attributes)


    def font_singular(self, cell_range: any, **attributes: any):
        column, row = Excel.convert_range(cell_range)

        attributes_string = Excel.font_attributes(**attributes)

        font = eval(f"Font({attributes_string})")
        self.active_sheet.cell(row = row, column = column).font = font

    
    def font_multiple(self, start_range: any, end_range: any, **attributes: any):
        start_column, start_row = Excel.convert_range(start_range)
        end_column, end_row = Excel.convert_range(end_range)

        attributes_string = Excel.font_attributes(**attributes)

        font = eval(f"Font({attributes_string})")
        for row in range(start_row, end_row + 1):
            for column in range(start_column, end_column + 1):
                self.active_sheet.cell(row = row, column = column).font = font

    
    #endregion Font


    #region Border
    def border_attributes(**attributes: any):
        list_of_attributes = []
        if("style" in attributes):
            if(type(attributes.get("style")) == str):
                if(attributes.get("style").lower() == "none"):
                    temp_attribute = f"border_style=None"

                else:
                    border_style = attributes.get("style").replace(" ", "")
                    border_style = border_style[0].lower() + border_style[1:]
                    temp_attribute = f"border_style='{border_style}'"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Style data type needs to be a string")

        if("color" in attributes):
            if(type(attributes.get("color")) == str):
                border_color = attributes.get("color")
                temp_attribute = f"color='{border_color}'"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Color data type needs to be a string")

        return Excel.attributes_string(list_of_attributes)


    def border_set(side, attribute):
        if(type(side) == str):
            side = side.lower()
        
        else:
            raise TypeError("Side data type needs to be a string")

        if(side == "all"):
            border = Border(top = attribute, left = attribute, right = attribute, bottom = attribute)

        elif(side == "top"):
            border = Border(top = attribute)

        elif(side == "left"):
            border = Border(left = attribute)

        elif(side == "right"):
            border = Border(right = attribute)

        elif(side == "bottom"):
            border = Border(bottom = attribute)
        
        else:
            raise TypeError("Side value can only be all, top, left, right, bottom")

        return border


    def border_singular(self, cell_range: any, side: str, **attributes: any):
        column, row = Excel.convert_range(cell_range)
            
        attributes_string = Excel.border_attributes(**attributes)
        
        side = eval(f"Side({attributes_string})")
        border = Excel.border_set(side)
        self.active_sheet.cell(row = row, column = column).border = border

    
    def border_multiple(self, start_range: any, end_range: any, side: str, **attributes: any):
        start_column, start_row = Excel.convert_range(start_range)
        end_column, end_row = Excel.convert_range(end_range)
            
        attributes_string = Excel.border_attributes(**attributes)
        
        attribute = eval(f"Side({attributes_string})")
        border = Excel.border_set(side, attribute)

        for row in range(start_row, end_row + 1):
            for column in range(start_column, end_column + 1):
                self.active_sheet.cell(row = row, column = column).border = border
    
    #endregion


    #region Alignment
    def alignment_attributes(**attributes: any):
        list_of_attributes = []
        if("horizontal" in attributes):
            if(type(attributes.get("horizontal")) == str):
                horizontal_type = attributes.get("horizontal")
                horizontal_type = (horizontal_type[0].lower() + horizontal_type[1:]).replace(' ', '')
                temp_attribute = f"horizontal='{horizontal_type}'"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Horizontal data type needs to be a string")

        if("vertical" in attributes):
            if(type(attributes.get("vertical")) == str):
                vertical_type = attributes.get("vertical")
                vertical_type = (vertical_type[0].lower() + vertical_type[1:]).replace(' ', '')
                temp_attribute = f"vertical='{vertical_type}'"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Vertical data type needs to be a string")

        if("rotation" in attributes):
            if(type(attributes.get("rotation")) in (str, int)):
                rotate_degree = int(attributes.get("rotation"))
                temp_attribute = f"text_rotation={rotate_degree}"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Rotation data type needs to be a string or an integer")

        if("indent" in attributes):
            if(type(attributes.get("indent")) in (str, int)):
                indent_value = int(attributes.get("indent"))
                temp_attribute = f"indent={indent_value}"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Indent data type needs to be a string or an integer")

        if("wrap" in attributes):
            if(type(attributes.get("wrap")) == bool):
                is_wrap = attributes.get("wrap")
                temp_attribute = f"wrap_text={is_wrap}"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Wrap data type needs to be a boolean")

        if("shrink" in attributes):
            if(type(attributes.get("shrink")) == bool):
                is_shrink = attributes.get("shrink")
                temp_attribute = f"shrink_to_fit={is_shrink}"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Shrink data type needs to be a boolean")

        return Excel.attributes_string(list_of_attributes)
    

    def alignment_singular(self, cell_range: any, **attributes: any):
        column, row = Excel.convert_range(cell_range)

        attributes_string = Excel.alignment_attributes(**attributes)
        
        alignment = eval(f"Alignment({attributes_string})")
        self.active_sheet.cell(row = row, column = column).alignment = alignment

    
    def alignment_multiple(self, start_range: any, end_range: any, **attributes: any):
        start_column, start_row = Excel.convert_range(start_range)
        end_column, end_row = Excel.convert_range(end_range)

        attributes_string = Excel.alignment_attributes(**attributes)

        alignment = eval(f"Alignment({attributes_string})")
        for row in range(start_row, end_row + 1):
            for column in range(start_column, end_column + 1):
                self.active_sheet.cell(row = row, column = column).alignment = alignment
                
    
    #endregion


class Database():
    def get_cluster(mongoDBURI):
        cluster = MongoClient(mongoDBURI)
        return cluster


    def get_database(mongoDBURI, database_name):
        db = Database.get_cluster(mongoDBURI)[database_name]
        return db

    
    def get_collection(mongoDBURI, database_name, collection_name):
        collection = Database.get_database(mongoDBURI, database_name)[collection_name]
        return collection


class Utility():
    def write_json(data, path):
        json_object = json.dumps(data, indent = 4)

        with open(path, "w") as outfile:
            outfile.write(json_object)


    def convert_to_dict(count, value, attribute, percentage_cell=[]):
        for i in range(len(percentage_cell)):
            if(type(value[percentage_cell[i]]) in (int, float)): 
                value[percentage_cell[i]] = round(value[percentage_cell[i]] * 100)

            if(type(value[percentage_cell[i]]) == str): 
                if(value[percentage_cell[i]] == "#REF!"):
                    value[percentage_cell[i]] = None


        temp_data_dictionary = {
            "id": count + 1
        }

        for i in range(len(attribute)):
            temp_data_dictionary[attribute[i]] = value[i]


        return temp_data_dictionary


    def update_data(mongoDBURI, database_name):
        excel_path = Main.excel_path

        excel_last_modified = os.path.getmtime(excel_path)
        translated_excel_last_modified = datetime.datetime.fromtimestamp(excel_last_modified).strftime("%y-%m-%d-%H-%M-%S")

        collection_name = "utilities"
        utilities_collection = Database.get_collection(mongoDBURI, database_name, collection_name)

        update_dictionary = {
            "id": 1,
            "last_modified": translated_excel_last_modified,
            "last_runned": datetime.datetime.now().strftime("%y-%m-%d-%H-%M-%S")
        }

        utilities_collection.replace_one({"id": 1}, update_dictionary, upsert=True)


class File():
    def create_file(mongoDBURI, file_path, data):
        system_path = os.getcwd()
        excel_client_dispatch = client.Dispatch("Excel.Application")

        current_datetime = datetime.datetime.now().strftime("%y-%m-%d-%H-%M-%S")
        excel_file_path = f"{current_datetime}.xlsx"
        pdf_file_path = f"{current_datetime}.pdf"

        if(Main.production_status == "Production"):
            database_phase = "Pro"

        elif(Main.production_status == "Development"):
            database_phase = "Dev"

        for division in data[0].get("divisi"):

            excel_folder_path = f"excel/{division.get('divisi').lower()}"
            full_excel_file_path = f"{file_path}/{excel_folder_path}/{excel_file_path}"
            system_excel_path = os.path.join(system_path, full_excel_file_path)

            pdf_folder_path = f"pdf/{division.get('divisi').lower()}"
            full_pdf_folder_path = f"{file_path}/{pdf_folder_path}/{pdf_file_path}"
            system_pdf_path = os.path.join(system_path, full_pdf_folder_path)

            os.makedirs(f"{file_path}/{excel_folder_path}", exist_ok=True)
            os.makedirs(f"{file_path}/{pdf_folder_path}", exist_ok=True)

            print(f"Membuat    : File {division.get('divisi')}")
            File.create_excel(file_path, division, full_excel_file_path)
            File.create_pdf(system_excel_path, system_pdf_path, excel_client_dispatch)

            print(f"Mengunggah : File {division.get('divisi')}")
            database_name = f"{database_phase}{division.get('divisi')}"
            File.upload_file(mongoDBURI, database_name, current_datetime, full_excel_file_path, full_pdf_folder_path)
            print(f"Selesai    : File {division.get('divisi')}")
            print()

        
        excel_client_dispatch.Application.Quit()


    def create_excel(file_path, division, full_excel_file_path):
        Excel.create_file(full_excel_file_path)

        wb_excel = Excel(full_excel_file_path, 1)
        wb_excel.write_value_singular("A1", f"Divisi: {division.get('divisi')}")
        wb_excel.write_value_multiple("A3", "C3", ["No.", "Sub Kegiatan", "Realisasi"])
        wb_excel.write_value_multiple("c4", "D4", ["Fisik", "Keuangan"])

        wb_excel.merge("A3", "A4")
        wb_excel.merge("B3", "B4")
        wb_excel.merge("C3", "D3")

        row_end_range = 4 + len(division.get('sub_kegiatan'))
        wb_excel.border_multiple("A3", [4, row_end_range], "all", style="thin")
        wb_excel.alignment_multiple("A3", [4, row_end_range], horizontal="center", vertical="center")
        wb_excel.alignment_multiple("B5", [2, row_end_range], horizontal="left", vertical="center", wrap=True)
        wb_excel.font_singular("A1", size=12, bold=True)
        wb_excel.font_multiple("A3", "D4", bold=True)

        for activity_count, activity in enumerate(division.get('sub_kegiatan')):
            activity_physical_value = None
            if(type(activity.get('fisik')) in (int, float)):
                activity_physical_value = f"{activity.get('fisik')}%"

            activity_finance_value = None
            if(type(activity.get('keuangan')) in (int, float)):
                activity_finance_value = f"{activity.get('keuangan')}%"

            activity_value = [
                activity_count + 1,
                activity.get("sub_kegiatan"),
                activity_physical_value,
                activity_finance_value
            ]

            wb_excel.write_value_multiple(
                [1, 5 + activity_count], [4, 5 + activity_count], activity_value)


        wb_excel.active_sheet.column_dimensions["A"].width = 5
        wb_excel.active_sheet.column_dimensions["B"].width = 60
        wb_excel.active_sheet.column_dimensions["C"].width = 11
        wb_excel.active_sheet.column_dimensions["D"].width = 11
        wb_excel.save()


    def create_pdf(system_excel_path, system_pdf_path, excel_client_dispatch):
        sheets = excel_client_dispatch.Workbooks.Open(system_excel_path)
        work_sheets = sheets.Worksheets[0]
    
        work_sheets.ExportAsFixedFormat(0, system_pdf_path)

        sheets.Close(True)


    def upload_file(mongoDBURI, database_name, file_name, excel_path, pdf_path):
        database = Database.get_database(mongoDBURI, database_name)
        
        excel_file = open(excel_path, "rb")
        pdf_file = open(pdf_path, "rb")

        excel_data = excel_file.read()
        pdf_data = pdf_file.read()

        fs = gridfs.GridFS(database)
        fs.put(excel_data, filename=f"{file_name}.xlsx")
        fs.put(pdf_data, filename=f"{file_name}.pdf")


class Main(Database, Utility, File):
    db_URI = None
    excel_path = None
    json_path = None
    file_path = None
    production_status = None

    def main():
        try:
            start_time = timeit.default_timer()

            mongoDBURI = Main.db_URI
            if(Main.production_status == "Development"):
                database_name = "Development"

            if(Main.production_status == "Production"):
                database_name = "DisnakerFinanceRecap"

            Main.get_data(mongoDBURI, database_name)

            # Main.show_data(mongoDBURI, database_name)

            Utility.update_data(mongoDBURI, database_name)

            stop_time = timeit.default_timer()
            elapsed_time = stop_time - start_time

            print(f"Waktu yang berlalu: {elapsed_time} detik")
            print()
            print("Anda bisa menutup programnya...")
            input()

        except:
            print("Program gagal untuk dijalankan...")
            input()


    def get_data(mongoDBURI, database_name):
        collection_name = "settings"

        settings_collection = Main.get_collection(mongoDBURI, database_name, collection_name)
        settings_data = list(settings_collection.find({}))
        
        excel_path = Main.excel_path
        wb_summary = Excel(excel_path, 1)

        division_array = Main.get_division(settings_data, wb_summary)

        recap_array = [
            {
                "id": 1,
                "tahun": 2022,
                "divisi": division_array
            }
        ]


        if(Main.production_status == "Development"):
            Main.write_json(recap_array, Main.json_path)
            data = json.load(open(Main.json_path))
            Main.update_data(mongoDBURI, database_name, data)

            Main.create_file(mongoDBURI, Main.file_path, data)            

        elif(Main.production_status == "Production"):
            Main.update_data(mongoDBURI, database_name, recap_array)

            Main.create_file(mongoDBURI, Main.file_path, recap_array)


    def get_division(settings_data, wb_summary):
        division_array = []
        for division_count, division_data in enumerate(settings_data):
            print(f"Memproses  : {division_data.get('name')}")
            activity_array = Main.get_activity(division_data, wb_summary)

            division_attribute = ["divisi", "sub_kegiatan"]
            division_value = [str(division_data.get("name")), activity_array]
            temp_division_dictionary = Main.convert_to_dict(division_count, division_value, division_attribute)

            division_array.append(temp_division_dictionary)
            print(f"Selesai    : {division_data.get('name')}")
            print()

        return division_array


    def get_activity(division_data, wb_summary):
        activity_array = []
        wb_summary.change_sheet(1)
        activity_value = wb_summary.get_value_multiple_2d(division_data.get("start_range"), division_data.get("end_range"))
        for activity_count, activity_data in enumerate(activity_value):
            detail_array = Main.get_detail(division_data, activity_count, wb_summary)

            activity_attribute = ["sub_kegiatan", "fisik", "keuangan", "detail"]
            activity_dict_value = [str(activity_data[0]), activity_data[1], activity_data[2], detail_array]
            activity_percentage_cell = [1, 2]
            temp_activity_dictionary = Main.convert_to_dict(activity_count, activity_dict_value, activity_attribute, activity_percentage_cell)

            activity_array.append(temp_activity_dictionary)


        return activity_array


    def get_detail(division_data, activity_count, wb_summary):
        detail_setting = division_data.get("detail")[activity_count]

        print(f"Memproses  : {detail_setting.get('id')} {detail_setting.get('active_sheet')} {detail_setting.get('sheet_name')} {detail_setting.get('start_range')} {detail_setting.get('end_range')} {detail_setting.get('attribute')}")

        detail_array = []
        if("null" not in detail_setting.get("sheet_name")):
            cell_range = [Excel.convert_range(detail_setting.get("start_range")), Excel.convert_range(detail_setting.get("end_range"))]

            cell_attribute_index = 0
            cell_attribute_count = 0

            detail_range = []
            expenses_range = []

            i = cell_range[0][1]
            while i < cell_range[1][1] + 2:
                if(cell_attribute_count == detail_setting.get("attribute")[cell_attribute_index]):  
                    combined_expenses_range = []

                    j = 0
                    while j < len(expenses_range):
                        temp_combined_expenses_range = expenses_range[j] + expenses_range[j+1] 
                        combined_expenses_range.append(temp_combined_expenses_range)

                        j += 2


                    temp_detail.append(combined_expenses_range)
                    detail_range.append(temp_detail)

                    expenses_range = []
                    cell_attribute_count = 0
                    cell_attribute_index += 1

                if(cell_attribute_count == 0):
                    temp_detail = [[cell_range[0][0], i], [cell_range[1][0], i]]

                    cell_attribute_count += 1
                    i += 1
                
                elif(cell_attribute_count != 0):
                    temp_expenses = [[cell_range[0][0], i], [cell_range[1][0], i+1]]
                    expenses_range.append(temp_expenses)

                    cell_attribute_count += 1
                    i += 2        

        
            wb_summary.change_sheet(detail_setting.get("sheet_name"))
            for i in range(len(detail_range)):
                expenses_array = []
                for j in range(len(detail_range[i][2])):
                    physical_value = wb_summary.get_value_multiple_2d(detail_range[i][2][j][0], detail_range[i][2][j][1])
                    finance_value = wb_summary.get_value_multiple_2d(detail_range[i][2][j][2], detail_range[i][2][j][3])

                    del physical_value[0][2:4]
                    del physical_value[1][2:4]
                    del finance_value[0][2:4]
                    del finance_value[1][2:4]

                    physical_monthly = []
                    for k in range(len(physical_value[0][2:14])):
                        temp_physical_monthly_dictionary = {
                            "id": k + 1,
                            "target": {
                                "total": physical_value[0][2:14][k],
                                "note": None
                            },
                            "realisasi": {
                                "total": physical_value[1][2:14][k],
                                "note": None
                            }
                        }

                        physical_monthly.append(temp_physical_monthly_dictionary)


                    finance_monthly = []
                    for k in range(len(finance_value[0][2:14])):
                        temp_finance_monthly_dictionary = {
                            "id": k + 1,
                            "target": {
                                "total": finance_value[0][2:14][k],
                                "note": None
                            },
                            "realisasi": {
                                "total": finance_value[1][2:14][k],
                                "note": None
                            }
                        }

                        finance_monthly.append(temp_finance_monthly_dictionary)


                    temp_expenses_dictionary = {
                        "id": j+1,
                        "biaya": str(physical_value[0][0]),
                        "fisik": {
                            "jumlah_fisik": str(physical_value[0][1]),
                            "jumlah_Kebutuhan_dana": physical_monthly
                        },
                        "keuangan": {
                            "jumlah_anggaran": str(finance_value[0][1]),
                            "jumlah_Kebutuhan_dana": finance_monthly
                        }
                    }

                    expenses_array.append(temp_expenses_dictionary)


                value = wb_summary.get_value_multiple(detail_range[i][0], detail_range[i][1])
                del value[2:4]

                monthly_total = []
                for monthly_value_index, monthly_value in enumerate(value[2:14]):
                    new_value_dict = {
                        "id": monthly_value_index + 1,
                        "total": monthly_value,
                        "note": None
                    }

                    monthly_total.append(new_value_dict)


                temp_detail_dictionary = {
                    "id": 1,
                    "rekening": str(value[0]),
                    "jumlah_fisik_anggaran": value[1],
                    "jumlah_Kebutuhan_dana": monthly_total,
                    "biaya": expenses_array
                }

                detail_array.append(temp_detail_dictionary)


        return detail_array


    def update_data(mongoDBURI, database_name, data):
        print("Mengunggah : Data")

        collection_name = "summary_recaps"
        summary_recaps_collection = Main.get_collection(mongoDBURI, database_name, collection_name)

        for i in range(len(data)):
            update_id = data[i].get("id")
            update_dictionary = {
                "id": update_id,
                "tahun": data[i].get("tahun"),
                "divisi": data[i].get("divisi")
            }

            summary_recaps_collection.replace_one({"id": update_id}, update_dictionary, upsert=True)


        print("Selesai    : Data")
        print()


    def show_data(mongoDBURI, database_name):
        collection_name = "summary_recaps"
        collection = Main.get_collection(mongoDBURI, database_name, collection_name)

        all_data = collection.find({})
        single_data = collection.find({"name": "Sekretariat"})

        print(all_data)
        print(single_data)


Main.main()
