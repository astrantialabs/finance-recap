import openpyxl

from openpyxl.styles import *

class Excel():
    def __init__(self, file_path: str, active_sheet: int = 1):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)

        wb_sheet = self.workbook.sheetnames

        self.active_sheet = self.workbook[wb_sheet[active_sheet - 1]]


    def create_file(file_path: str):
        wb = openpyxl.Workbook()
        wb.save(file_path)


    def save(self):
        self.workbook.save(self.file_path)


    def create_sheet(self, new_sheet_name: str):
        self.workbook.create_sheet(new_sheet_name)


    def change_sheet(self, active_sheet: int or str):
        if(type(active_sheet) == int):
            self.active_sheet = self.workbook[self.workbook.sheetnames[active_sheet - 1]]

        elif(type(active_sheet) == str):
            self.active_sheet = self.workbook[active_sheet]


    def change_sheet_name(self, old_sheet_name, new_sheet_name):
        self.workbook[old_sheet_name].title = new_sheet_name
    

    def delete_sheet(self, deleted_sheet: int or str):
        if(type(deleted_sheet) == int):
            del self.workbook[self.workbook.sheetnames[deleted_sheet - 1]]

        elif(type(deleted_sheet) == str):
            del self.workbook[deleted_sheet]


    def adjust_width(self, start_range: any, end_range: any, extra_width: int = 0, width_limit: int = 0):
        start_column, start_row = Excel.convert_range(start_range)
        end_column, end_row = Excel.convert_range(end_range)

        width_dict = {}
        for column in range(start_column, end_column + 1):
            temp_width_array = []
            cell_column_letter =  self.active_sheet.cell(row = 1, column = column).column_letter

            for row in range(start_row, end_row + 1):
                temp_value = self.active_sheet.cell(row = row, column = column).value

                if(type(temp_value) in (int, str)):
                    temp_width_array.append(len(str(temp_value)))


            max_width_value = max(temp_width_array)

            if(width_limit > 0):
                if(max_width_value > width_limit):
                    max_width_value = width_limit

                    self.alignment_multiple([column, start_row], [column, end_row], wrap = True)


            width_dict[cell_column_letter] = max_width_value

        
        for column, width in width_dict.items():
            self.active_sheet.column_dimensions[column].width = (width + 1 + extra_width)


    def check_range(range: any):
        if (type(range) not in (str, list)):
            raise TypeError("Range must be a type of string or list")

        elif (type(range) == str):
            if (range.isalpha() or range.isnumeric()):
                raise TypeError("Range string must be a combination of character and number")

        elif (type(range) == list):
            if (len(range) == 2):
                for i in range:
                    if type(i) not in (str, int):
                        raise TypeError("Range list can only have a type of string and integer for its values")

            else:
                raise TypeError("Range list can only have 2 values")
                

    def convert_range(range: any):
        Excel.check_range(range)

        if (type(range) == str):
            column = Excel.check_and_convert_string_value(''.join(x for x in range if not x.isdigit()))
            row = int(''.join(x for x in range if x.isdigit()))

        elif (type(range) == list):
            if (type(range[0]) == str):
                column = Excel.check_and_convert_string_value(range[0])
            
            elif (type(range[0]) == int):
                column = range[0]

            if (type(range[1]) == str):
                row = Excel.check_and_convert_string_value(range[1])
            
            elif (type(range[1]) == int):
                row = range[1]

        return column, row


    def check_and_convert_string_value(value: any):
        if(type(value) == str):
            value = [ord(x) - 96 for x in value.lower()]

            new_value = 0
            for i in range(len(value)):
                new_value += value[i] * 26**(len(value) - (i + 1))

        return new_value

    
    def attributes_string(list_of_attributes: any):
        attributes_string = ""
        for i, attribute in enumerate(list_of_attributes):
            if(i == 0):
                attributes_string += attribute
            
            else:
                attributes_string += f", {attribute}"

        return attributes_string


    def change_cell_number_format_singular(self, range: any, number_format: str):
        column, row = Excel.convert_range(range)
        self.active_sheet.cell(row = row, column = column).number_format = number_format


    def change_cell_number_format_multiple(self, start_range: any, end_range: any, number_format: str):
        start_column, start_row = Excel.convert_range(start_range)
        end_column, end_row = Excel.convert_range(end_range)

        for row in range(start_row, end_row + 1):
            for column in range(start_column, end_column + 1):
                self.active_sheet.cell(row = row, column = column).number_format = number_format


    #region Get
    def get_value_singular(self, range: any):
        column, row = Excel.convert_range(range)
        value = self.active_sheet.cell(row = row, column = column).value

        return value

        
    def get_value_multiple(self, start_range: any, end_range: any):
        start_column, start_row = Excel.convert_range(start_range)
        end_column, end_row = Excel.convert_range(end_range)
        
        value = []
        for row in range(start_row, end_row + 1):
            for column in range(start_column, end_column + 1):
                temp_value = self.active_sheet.cell(row = row, column = column).value
                value.append(temp_value)

        return value


    def get_value_multiple_2d(self, start_range: any, end_range: any):
        start_column, start_row = Excel.convert_range(start_range)
        end_column, end_row = Excel.convert_range(end_range)

        value_array = []
        for row in range(start_row, end_row + 1):
            temp_value_array = []
            for column in range(start_column, end_column + 1):
                temp_value = self.active_sheet.cell(row = row, column = column).value
                temp_value_array.append(temp_value)
            
            value_array.append(temp_value_array)

        return value_array

    #endregion Get


    #region Write
    def write_value_singular(self, range: any, value: any):
        if(type(value) == list):
            raise TypeError("Use write_value_multiple function if the value type is a list")

        column, row = Excel.convert_range(range)
            
        self.active_sheet.cell(row = row, column = column, value = value)


    def write_value_multiple(self, start_range: any, end_range: any, value: any):
        start_column, start_row = Excel.convert_range(start_range)
        end_column, end_row = Excel.convert_range(end_range)
        
        if(type(value) == list):
            for check_value in value:
                if(type(check_value) == list):
                    raise TypeError("Use write_value_multiple_2d function if the value is a 2D list")


            value_counter = 0
            for row in range(start_row, end_row + 1):
                for column in range(start_column, end_column + 1):
                    self.active_sheet.cell(row = row, column = column, value = value[value_counter])
                    value_counter += 1


        elif(type(value) in (str, int, bool, float)):
            for row in range(start_row, end_row + 1):
                for column in range(start_column, end_column + 1):
                    self.active_sheet.cell(row = row, column = column, value = value)       


    def write_value_multiple_2d(self, start_range: any, value: any):
        if(type(value) == list):
            value_is_valid = True
            for check_value in value:
                if(type(check_value) != list):
                    value_is_valid = False


            if(value_is_valid):
                start_column, start_row = Excel.convert_range(start_range)
                end_column, end_row = start_column + len(value[0]), start_row + len(value)

                for x, row in enumerate(range(start_row, end_row)):
                    for y, column in enumerate(range(start_column, end_column)):
                        self.active_sheet.cell(row = row, column = column, value = value[x][y])


            elif(not value_is_valid):
                raise TypeError("Value must be a 2D list")
        
        else:
            raise TypeError("Value must be a 2D list")

    #endregion Write


    #region Merge & Unmerge
    def merge(self, start_range: any, end_range: any):
        start_column, start_row = Excel.convert_range(start_range)
        end_column, end_row = Excel.convert_range(end_range)

        self.active_sheet.merge_cells(start_row = start_row, start_column = start_column, end_row = end_row, end_column = end_column)

    
    def unmerge(self, start_range: any, end_range: any):
        start_column, start_row = Excel.convert_range(start_range)
        end_column, end_row = Excel.convert_range(end_range)

        self.active_sheet.unmerge_cells(start_row = start_row, start_column = start_column, end_row = end_row, end_column = end_column)

    #endregion Merge & Unmerge


    #region Font
    def font_attributes(**attributes: any):
        list_of_attributes = []
        if("font" in attributes):
            if(type(attributes.get("font")) == str):
                font_name = attributes.get("font")
                temp_attribute = f"name='{font_name}'"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Font data type needs to be a string")

        if("size" in attributes):
            if(type(attributes.get("size")) in (str, int)):
                font_size = int(attributes.get("size"))
                temp_attribute = f"size={font_size}"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Size data type needs to be a string or an integer")

        if("color" in attributes):
            if(type(attributes.get("color")) == str):
                color_name = attributes.get("color")
                temp_attribute = f"color='{color_name}'"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Color data type needs to be a string")

        if("underline" in attributes):
            if(type(attributes.get("underline")) == str):
                underline_name = attributes.get("underline")
                underline_name = (underline_name[0].lower() + underline_name[1:]).replace(" ", "")
                temp_attribute = f"underline='{underline_name}'"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Underline data type needs to be a string")

        if("bold" in attributes):
            if(type(attributes.get("bold")) == bool):
                is_bold = attributes.get("bold")
                temp_attribute = f"bold={is_bold}"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Bold data type needs to be a boolean")

        if("italic" in attributes):
            if(type(attributes.get("italic")) == bool):
                is_italic = attributes.get("italic")
                temp_attribute = f"italic={is_italic}"

                list_of_attributes.append(temp_attribute)
            
            else:
                raise TypeError("Italic data type needs to be a boolean")

        if("strike" in attributes):
            if(type(attributes.get("strike")) == bool):
                is_strike = attributes.get("strike")
                temp_attribute = f"strike={is_strike}"

                list_of_attributes.append(temp_attribute)
        
            else:
                raise TypeError("Strike data type needs to be a boolean")

        return Excel.attributes_string(list_of_attributes)


    def font_singular(self, cell_range: any, **attributes: any):
        column, row = Excel.convert_range(cell_range)

        attributes_string = Excel.font_attributes(**attributes)

        font = eval(f"Font({attributes_string})")
        self.active_sheet.cell(row = row, column = column).font = font

    
    def font_multiple(self, start_range: any, end_range: any, **attributes: any):
        start_column, start_row = Excel.convert_range(start_range)
        end_column, end_row = Excel.convert_range(end_range)

        attributes_string = Excel.font_attributes(**attributes)

        font = eval(f"Font({attributes_string})")
        for row in range(start_row, end_row + 1):
            for column in range(start_column, end_column + 1):
                self.active_sheet.cell(row = row, column = column).font = font

    
    #endregion Font


    #region Fill
    def fill_attributes(**attributes: str):
        list_of_attributes = []

        if("type" in attributes):
            if(type(attributes.get("type")) == str):
                fill_type = attributes.get("type")
                fill_type = (fill_type[0].lower() + fill_type[1:]).replace(' ', '')
                temp_attribute = f"fill_type='{fill_type}'"

                list_of_attributes.append(temp_attribute)

            elif(attributes.get("type") == None):
                temp_attribute = f"fill_type=None"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Type data type needs to be a string or None")

        if("main_color" in attributes):
            if(type(attributes.get("main_color")) == str):
                fill_color = attributes.get("main_color")

                temp_attribute = f"start_color='{fill_color}'"
                
                list_of_attributes.append(temp_attribute)

        if("second_color" in attributes):
            if(type(attributes.get("second_color")) == str):
                fill_color = attributes.get("second_color")
                
                temp_attribute = f"end_color='{fill_color}'"
                    
                list_of_attributes.append(temp_attribute)

        return Excel.attributes_string(list_of_attributes)


    def shade_attributes(**attributes: str):
        list_of_attributes = []

        if("shade" in attributes):
            if(type(attributes.get("shade")) != bool):
                raise TypeError("Shade data type needs to be a boolean")

        if("type" in attributes):
            if(type(attributes.get("type")) == str):
                fill_type = attributes.get("type")
                fill_type = (fill_type[0].lower() + fill_type[1:]).replace(' ', '')
                temp_attribute = f"fill_type='{fill_type}'"

                list_of_attributes.append(temp_attribute)

            elif(attributes.get("type") == None):
                temp_attribute = f"fill_type=None"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Type data type needs to be a string or None")

        if("main_color" in attributes):
            if(type(attributes.get("main_color")) == str):
                fill_color = attributes.get("main_color")
                temp_attribute = f"end_color='{fill_color}'"
                
                list_of_attributes.append(temp_attribute)

        if("second_color" in attributes):
            if(type(attributes.get("second_color")) == str):
                fill_color = attributes.get("second_color")
                temp_attribute = f"start_color='{fill_color}'"
                    
                list_of_attributes.append(temp_attribute)

        return Excel.attributes_string(list_of_attributes)


    def fill_singular(self, cell_range: any, **attributes: any):
        column, row = Excel.convert_range(cell_range)
            
        attributes_string = Excel.fill_attributes(**attributes)

        self.active_sheet.cell(row = row, column = column).fill = eval(f"PatternFill({attributes_string})")

    
    def fill_multiple(self, start_range: any, end_range: any, **attributes: any):
        start_column, start_row = Excel.convert_range(start_range)
        end_column, end_row = Excel.convert_range(end_range)
            
        main_attributes_string = Excel.fill_attributes(**attributes)

        shade = False
        if("shade" in attributes):
            shade = attributes.get("shade")

        if(shade):    
            second_attributes_string = Excel.shade_attributes(**attributes)

        main_pattern_fill = eval(f"PatternFill({main_attributes_string})")
        second_pattern_fill = eval(f"PatternFill({second_attributes_string})")
        for row in range(start_row, end_row + 1):
            for column in range(start_column, end_column + 1):
                self.active_sheet.cell(row = row, column = column).fill = main_pattern_fill

                if(shade and column % 2 == 0):
                    self.active_sheet.cell(row = row, column = column).fill = second_pattern_fill
    

    #endregion Fill


    #region Border
    def border_attributes(**attributes: any):
        list_of_attributes = []
        if("style" in attributes):
            if(type(attributes.get("style")) == str):
                if(attributes.get("style").lower() == "none"):
                    temp_attribute = f"border_style=None"

                else:
                    border_style = attributes.get("style").replace(" ", "")
                    border_style = border_style[0].lower() + border_style[1:]
                    temp_attribute = f"border_style='{border_style}'"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Style data type needs to be a string")

        if("color" in attributes):
            if(type(attributes.get("color")) == str):
                border_color = attributes.get("color")
                temp_attribute = f"color='{border_color}'"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Color data type needs to be a string")

        return Excel.attributes_string(list_of_attributes)


    def border_set(side, attribute):
        if(type(side) == str):
            side = side.lower()
        
        else:
            raise TypeError("Side data type needs to be a string")

        if(side == "all"):
            border = Border(top = attribute, left = attribute, right = attribute, bottom = attribute)

        elif(side == "top"):
            border = Border(top = attribute)

        elif(side == "left"):
            border = Border(left = attribute)

        elif(side == "right"):
            border = Border(right = attribute)

        elif(side == "bottom"):
            border = Border(bottom = attribute)
        
        else:
            raise TypeError("Side value can only be all, top, left, right, bottom")

        return border


    def border_singular(self, cell_range: any, side: str, **attributes: any):
        column, row = Excel.convert_range(cell_range)
            
        attributes_string = Excel.border_attributes(**attributes)
        
        side = eval(f"Side({attributes_string})")
        border = Excel.border_set(side)
        self.active_sheet.cell(row = row, column = column).border = border

    
    def border_multiple(self, start_range: any, end_range: any, side: str, **attributes: any):
        start_column, start_row = Excel.convert_range(start_range)
        end_column, end_row = Excel.convert_range(end_range)
            
        attributes_string = Excel.border_attributes(**attributes)
        
        attribute = eval(f"Side({attributes_string})")
        border = Excel.border_set(side, attribute)

        for row in range(start_row, end_row + 1):
            for column in range(start_column, end_column + 1):
                self.active_sheet.cell(row = row, column = column).border = border
    
    #endregion


    #region Alignment
    def alignment_attributes(**attributes: any):
        list_of_attributes = []
        if("horizontal" in attributes):
            if(type(attributes.get("horizontal")) == str):
                horizontal_type = attributes.get("horizontal")
                horizontal_type = (horizontal_type[0].lower() + horizontal_type[1:]).replace(' ', '')
                temp_attribute = f"horizontal='{horizontal_type}'"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Horizontal data type needs to be a string")

        if("vertical" in attributes):
            if(type(attributes.get("vertical")) == str):
                vertical_type = attributes.get("vertical")
                vertical_type = (vertical_type[0].lower() + vertical_type[1:]).replace(' ', '')
                temp_attribute = f"vertical='{vertical_type}'"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Vertical data type needs to be a string")

        if("rotation" in attributes):
            if(type(attributes.get("rotation")) in (str, int)):
                rotate_degree = int(attributes.get("rotation"))
                temp_attribute = f"text_rotation={rotate_degree}"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Rotation data type needs to be a string or an integer")

        if("indent" in attributes):
            if(type(attributes.get("indent")) in (str, int)):
                indent_value = int(attributes.get("indent"))
                temp_attribute = f"indent={indent_value}"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Indent data type needs to be a string or an integer")

        if("wrap" in attributes):
            if(type(attributes.get("wrap")) == bool):
                is_wrap = attributes.get("wrap")
                temp_attribute = f"wrap_text={is_wrap}"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Wrap data type needs to be a boolean")

        if("shrink" in attributes):
            if(type(attributes.get("shrink")) == bool):
                is_shrink = attributes.get("shrink")
                temp_attribute = f"shrink_to_fit={is_shrink}"

                list_of_attributes.append(temp_attribute)

            else:
                raise TypeError("Shrink data type needs to be a boolean")

        return Excel.attributes_string(list_of_attributes)
    

    def alignment_singular(self, cell_range: any, **attributes: any):
        column, row = Excel.convert_range(cell_range)

        attributes_string = Excel.alignment_attributes(**attributes)
        
        alignment = eval(f"Alignment({attributes_string})")
        self.active_sheet.cell(row = row, column = column).alignment = alignment

    
    def alignment_multiple(self, start_range: any, end_range: any, **attributes: any):
        start_column, start_row = Excel.convert_range(start_range)
        end_column, end_row = Excel.convert_range(end_range)

        attributes_string = Excel.alignment_attributes(**attributes)

        alignment = eval(f"Alignment({attributes_string})")
        for row in range(start_row, end_row + 1):
            for column in range(start_column, end_column + 1):
                self.active_sheet.cell(row = row, column = column).alignment = alignment
                
    
    #endregion


    #region Function
    def summary(self, start_range:any, end_range:any):
        array_of_value = self.get_value_multiple(start_range, end_range)

        sum_value = 0
        for value in array_of_value:
            if (type(value) == int):
                sum_value += value

            elif (type(value) == str):
                if(value.isnumeric()):
                    sum_value += int(value)

        return sum_value


    def count(self, start_range:any, end_range:any):
        value_array = self.get_value_multiple(start_range, end_range)
        
        value = 0
        for i in value_array:
            if(type(i) in (int, float)):
                value += 1

        return value


    def count_a(self, start_range:any, end_range:any):
        value_array = self.get_value_multiple(start_range, end_range)

        value = 0
        for i in value_array:
            if(type(i) in (str, int, float, bool)):
                value += 1

        return value


    def count_blank(self, start_range:any, end_range:any):
        value_array = self.get_value_multiple(start_range, end_range)

        value = 0
        for i in value_array:
            if(type(i) not in (str, int, float, bool)):
                value += 1

        return value


    def average(self, start_range:any, end_range:any):
        total = self.summary(start_range, end_range)
        count = self.count(start_range, end_range)

        value = total / count

        return value


    def excel_if(self, range1: any, logic: str, range2: any, return1: any, return2: any): 
        value1 = self.get_value_singular(range1)
        value2 = self.get_value_singular(range2)

        if(logic == "="):
            if(value1 == value2):
                is_true = True
            
            else:
                is_true = False

        elif(logic == "!="):
            if(value1 != value2):
                is_true = True
            
            else:
                is_true = False

        elif(logic == ">"):
            if(value1 > value2):
                is_true = True
            
            else:
                is_true = False

        elif(logic == "<"):
            if(value1 < value2):
                is_true = True
            
            else:
                is_true = False

        elif(logic == ">="):
            if(value1 >= value2):
                is_true = True
            
            else:
                is_true = False
        
        elif(logic == "<="):
            if(value1 <= value2):
                is_true = True
            
            else:
                is_true = False
  
        if(is_true):
            return return1
    
        elif(not is_true):
            return return2


    def summary_if(self, start_criteria_range: any, end_criteria_range: any, criteria: str, sum_start_range: any, sum_end_range: any): 
        criteria_array = self.get_value_multiple(start_criteria_range, end_criteria_range)
        sum_array = self.get_value_multiple(sum_start_range, sum_end_range)

        if(len(criteria_array) == len(sum_array)):
            sum_value = 0
            for i in range(len(criteria_array)):
                if criteria_array[i] == criteria:
                    sum_value += sum_array[i]

            return sum_value


    def count_if(self, start_range: any, end_range: any, criteria: str): 
        value_array = self.get_value_multiple(start_range, end_range)

        value = 0
        for i in value_array:
                if(i == criteria):
                    value += 1

        return value

    
    def average_if(self, start_criteria_range: any, end_criteria_range: any, criteria: str, sum_start_range: any, sum_end_range: any):
        total_value = self.summary_if(start_criteria_range, end_criteria_range, criteria, sum_start_range, sum_end_range)
        count_value = self.count_if(start_criteria_range, end_criteria_range, criteria)

        return total_value / count_value


    def excel_max(self, start_range: any, end_range: any):
        max_value = max(self.get_value_multiple(start_range, end_range))

        return max_value


    def excel_min(self, start_range: any, end_range: any):
        min_value = min(self.get_value_multiple(start_range, end_range))

        return min_value

    #endregion
